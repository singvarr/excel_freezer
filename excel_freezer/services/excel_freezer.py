from os import path
from copy import copy
from datetime import datetime
from typing import Optional
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.cell_range import CellRange

from excel_freezer.models.freeze_table_configuration import (
    FreezeSheetSettings,
    FreezeTableConfiguration,
)


class ExcelFreezer:
    _DEFAULT_DATE_FORMAT = "%Y-%m-%d"

    def __init__(self, configuration: Optional[FreezeTableConfiguration]):
        self._configuration = configuration

    @staticmethod
    def get_default_output_file_name(source: Path) -> str:
        today = datetime.now().strftime(ExcelFreezer._DEFAULT_DATE_FORMAT)
        file_name = path.splitext(path.basename(source))[0]

        return f"{today}_{file_name}.xlsx"

    def _check_is_cell_formula_preserved(
        self,
        cell: Cell,
        sheet_settings: Optional[FreezeSheetSettings],
    ) -> bool:
        if not sheet_settings:
            return False

        return any(
            cell.coordinate in CellRange(cell_range)
            for cell_range in sheet_settings.cells_with_preserved_formulae
        )

    def run(self):
        wb_data = load_workbook(self._configuration.source, data_only=True)
        wb_styles = load_workbook(self._configuration.source, data_only=False)

        for sheet_name in wb_styles.sheetnames:
            sheet_settings = next(
                (
                    settings
                    for settings in self._configuration.sheet_settings
                    if settings.name == sheet_name
                ),
                None,
            )

            if sheet_settings and not sheet_settings.preserve_sheet:
                continue

            ws_style = wb_styles[sheet_name]
            ws_data = wb_data[sheet_name]

            for col_name, col_dim in ws_style.column_dimensions.items():
                ws_data.column_dimensions[col_name].width = col_dim.width
            for row_index, row_dim in ws_style.row_dimensions.items():
                ws_data.row_dimensions[row_index].height = row_dim.height

            for row in ws_style.iter_rows():
                for cell in row:
                    is_formula_preserved = self._check_is_cell_formula_preserved(
                        cell=cell,
                        sheet_settings=sheet_settings,
                    )

                    new_cell = ws_data.cell(row=cell.row, column=cell.column)

                    if is_formula_preserved:
                        new_cell.value = cell.value

                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = copy(cell.number_format)
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)

        wb_data._external_links = []
        wb_data.save(self._configuration.destination)
