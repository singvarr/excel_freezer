from datetime import datetime
from os import path
from copy import copy

import openpyxl

from excel_freezer.services.dialog_manager import DialogManager


def freeze_excel_values():
    dialog_manager = DialogManager()

    input_path = dialog_manager.request_table_to_freeze()

    if not input_path:
        return

    today = datetime.now().strftime("%d.%m.%Y")
    file_name = path.splitext(path.basename(input_path))[0]
    default_name = f"{today}_{file_name}.xlsx"

    output_path = dialog_manager.request_path_for_save(default_file_name=default_name)

    if not output_path:
        return

    try:
        wb_data = openpyxl.load_workbook(input_path, data_only=True)
        wb_styles = openpyxl.load_workbook(input_path, data_only=False)

        for sheet_name in wb_styles.sheetnames:
            ws_style = wb_styles[sheet_name]
            ws_data = wb_data[sheet_name]

            for col_name, col_dim in ws_style.column_dimensions.items():
                ws_data.column_dimensions[col_name].width = col_dim.width
            for row_index, row_dim in ws_style.row_dimensions.items():
                ws_data.row_dimensions[row_index].height = row_dim.height

            for row in ws_style.iter_rows():
                for cell in row:
                    new_cell = ws_data.cell(row=cell.row, column=cell.column)
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = copy(cell.number_format)
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)

        wb_data.save(output_path)

        dialog_manager.show_success_message(message=f"Готово!\nФайл збережено як:\n{output_path}")
    except Exception as error:
        dialog_manager.show_error_message(error=error)


if __name__ == "__main__":
    freeze_excel_values()